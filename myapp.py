from kivy.app import App
import math
from kivy.uix.label import Label
from kivy.core.window import Window
from kivy.uix.button import Button
from kivy.uix.boxlayout import BoxLayout
from kivy.uix.floatlayout import FloatLayout
from kivy.properties import ObjectProperty
from kivy.uix.gridlayout import GridLayout
from kivy.lang import Builder
from openpyxl import Workbook, load_workbook
from kivy.weakproxy import WeakProxy
import os

Builder.load_file('My.kv')

Window.size = (1920, 1080)
Window.clearcolor = (106/255, 106/255, 106/255)
Window.title = 'Приложение'


class Container(GridLayout):

    def calculate(self):

        wb = load_workbook("data.xlsx", read_only=True)
        ws = wb.active
        wb1 = load_workbook("data1.xlsx")
        ws1 = wb1.active

        a023 = (ws['F3'].value)
        a024 = (ws['H3'].value)
        a025 = (ws['G3'].value)
        m003 = (ws['D3'].value)
        c102 = (ws['X3'].value)
        t = (ws['N3'].value)

        a034 = a024 + ((c102 * 1000)/m003)
        ccu0 = a023 - ((a034 - a023) * 64.5 / 65.5)
        ccd0 = a025 - ((a034 - a025) * 112.4 / 65.5)

        ws1['B3'] = "%.2f" % ccu0
        ws1['C3'] = "%.2f" % ccd0
        wb1.save('data1.xlsx')

        self.test.text = str("%.2f" % ccd0)
        self.test1.text = str("%.2f" % ccu0)

    def test(self):
        os.startfile(r'data1.xlsx')


class MyApp(App):

    def build(self):



        return Container()

if __name__ == "__main__":
    MyApp().run()
